# -*- coding: UTF-8 -*-

import os
import collections
import platform 
import re
_version = platform.python_version()
import json
from datetime import date
import GTime



try:
	import openpyxl
except:
	try:
		if(int(_version[:1])<3):
			os.system("pip install openpyxl")
		else:
			os.system("pip3 install openpyxl")

		import openpyxl
	except:
		print '请先安装pip,openpyxl'
		exit(1)

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from copy import copy


timeList = GTime.outTime()


question = collections.OrderedDict()
question['name'] = ''
question['companyName'] = ''
question['position'] = ''

if os.path.exists('./setting.txt'):
	f = open('./setting.txt', 'r')
	for line in f.readlines():
		item = line.strip().split(':')
		if(item[0] in question):
			question[item[0]] = item[1];
	f.close()


questionMap = {
	'name':'请输入你的名字:',
	'companyName':'请输入你的公司名:',
	'position':'请输入你的职位:'
}

rules = {}

try:
	with open("./rules.txt","r") as fs:
		content = fs.read()  
		rules = json.loads(content)
except:
	print '请确认当前文件夹下有rules文件'
	exit(1)



def set():
	for i in question:
		ans = raw_input(questionMap[i])
		question[i] = ans


def save():
	with open("./setting.txt","w") as fs:
		for i in question:
			fs.write("%s:%s\n" % (i,question[i]))

def checkSet():
	for i in question:
		if(question[i].strip() == ''):
			return 'error'

def getTheWeek():
	if date.today().day<10:
		return 1
	elif date.today().day<21:
		return 2
	elif date.today().day<28:
		return 3
	else:
		return 4

def outFileName():
	return '%s-%s月第%s周周报' % (question['name'],date.today().month,getTheWeek())


def replace(pos,str):
	rep = str[pos[0]+1:pos[1]-1]
	if rep in question:
		return '%s%s%s' % (str[:pos[0]],question[rep],str[pos[1]:])
	elif rep in rules:
		if rep == 'summer':
			return '本周工作总结:'+raw_input(rules[rep].encode('utf-8')+':')
		else:
			return raw_input(rules[rep].encode('utf-8')+':')
	elif rep in timeList:
		try:
			return '%s%s%s' % (str[:pos[0]],timeList[rep],str[pos[1]:])
		except:
			return '%s%s%s' % (str[:pos[0]],timeList[rep].decode('utf-8'),str[pos[1]:])
	elif '.' in rep:
		index = rep.index('.')
		prefix = rep[:index]
		if prefix in timeList:
			return timeList[prefix][rep[index+1:]]
	else:
		return '{'+rep+'}'

def search(str):
	result = re.search('{[\.|\w]+}',str)
	if result == None:
		return str
	pos = result.span()
	afterRep = replace(pos,str)
	return search(afterRep)
 


def readFile():
	try:
		wb = load_workbook('./template.xlsx')
	except:
		print '请确认当前文件夹下有模板文件'
		exit(1)
	oldSheet = wb['Sheet1']
	thin = Side(border_style="thin", color="000000")
	fill = PatternFill("solid", fgColor="ffffff")
	border = Border(top=thin, left=thin, right=thin, bottom=thin)
	
	for row in oldSheet.rows:
		for cell in row:
			cell.border = border
			cell.fill = fill
			if cell.value == '' or cell.value == None:
				continue
			if isinstance(cell.value,basestring) == True:
				# print 'run'
				if '{' in cell.value:
					cell.value = search(cell.value)


	print '文档生成中..'
	wb.save('output/%s.xlsx' % (outFileName()))
	print '文档生成完毕,请在 output 查看'


def run():
	firstSelect = raw_input('请选择要做的事(s:设置, g:开始生成周报):')
	if(firstSelect == 's'):
		set()
		print '设置完毕'
		save()
		return run()
	elif(firstSelect == 'g'):
		if(checkSet() == 'error'):
			print '请先设置个人信息'
			return run() # TODO 后续优化 只需输入未填写的信息
		else:
			print '模板读取中..'
			readFile()
	else:
		print '请重新选择'
		return run()

run()